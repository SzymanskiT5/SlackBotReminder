import datetime

from openpyxl import load_workbook
import re
from openpyxl.worksheet.worksheet import Worksheet
from constans import DATE_REGEX, MONTH_REGEX
from main import db
from models import Student


class SheetScraper:
    def __init__(self, localization):
        self.localization = localization
        self.ws = self.open_workbook()
        self.list_of_students = []

    def open_workbook(self) -> Worksheet:
        wb = load_workbook(self.localization)
        ws = wb.active
        return ws

    def iterate_rows(self) -> None:
        for idx, row in enumerate(self.ws.iter_rows(min_row=7)):
            start_index = 5
            if re.match(DATE_REGEX, str(row[start_index].value)):
                name = str(row[1].value)
                print(name)
                last_meeting_index = self.find_last_meeting(row, start_index)
                check_date = self.check_date_and_month(str(row[last_meeting_index].value))
                print(check_date)
                curr_month = self.check_neighbouring_meetings(row, last_meeting_index, idx)
                if curr_month and check_date:
                    self.check_if_student_exists(name, curr_month)

    def find_last_meeting(self, row, start_index) -> int:
        diff = 4
        while row[start_index].value != None:
            start_index += diff
        start_index -= diff
        last_fourth_meeting = row[start_index].value
        print(last_fourth_meeting)
        return start_index

    def check_date_and_month(self, meeting_date):
        """In Sheet is used dd-mm or d-mm format, this method avoids mistakes when someone
         adds meetings before they happened """
        dot_index = meeting_date.find('.')
        meeting_day = int(meeting_date[:dot_index])
        meeting_month = int(meeting_date[dot_index + 1:])
        today_day = datetime.date.today().day
        today_month = datetime.date.today().month
        if today_month == 1 and meeting_month == 12 or today_month > meeting_month or \
                (today_month == meeting_month and today_day > meeting_day):
            return True

        return False


    def check_neighbouring_meetings(self, row, index, idx) -> str:
        if not re.match(DATE_REGEX, str(row[index + 1].value)) and re.match(DATE_REGEX, str(row[index - 1].value)):
            while not re.search(MONTH_REGEX, str(self.ws.cell(row=idx + 7, column=index - 1).value)):
                idx -= 1
            curr_month = str(self.ws.cell(row=idx + 7, column=index - 1).value)
            print(curr_month)
            return curr_month

    def check_if_student_exists(self, name, curr_month) -> None:
        student = db.session.query(Student).filter_by(name=name).first()

        if not student:
            student = Student(name=name, last_month=curr_month)
            db.session.add(student)
            db.session.commit()
            self.list_of_students.append(name)

        elif student.last_month != curr_month:
            self.list_of_students.append(name)
